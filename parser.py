# -*- coding: utf-8 -*-
"""
Разбор журнала смывов: 13 вкладок Excel с разными шапками -> одна чистая таблица.
Чинит объединённые даты, переносы строк в названиях, типизирует значения,
находит несоответствия и превышения норматива КМАФАнМ (1x10^3 КОЕ/см²).
"""
import re
import pandas as pd
import openpyxl

# Показатели и синонимы в шапках листов
INDICATORS = {
    "КМАФАнМ": ["кмафанм"],
    "БГКП": ["бгкп"],
    "Proteus": ["proteus"],
    "Salmonella": ["salmonella"],
    "Listeria": ["monocytogenes", "listeria"],
    "Staph": ["staph"],
    "Плесень": ["плесень"],
    "Дрожжи": ["дрожжи"],
}
KMA_LIMIT = 1000.0  # КОЕ/см², норматив КМАФАнМ

# Листы, которые не являются журналами цехов
SKIP_SHEETS = {"лист2"}


def _norm(v):
    return "" if v is None else str(v).strip()


def _clean_loc(s):
    """Убрать переносы строк и лишние пробелы в названии точки отбора."""
    return re.sub(r"\s+", " ", s.replace("\n", " ")).strip()


def parse_kma(raw):
    """
    Разобрать значение КМАФАнМ в число КОЕ/см².
    'менее 1,5 х 10*2' -> 150 (порог обнаружения), '6,4 х 10*3' -> 6400.
    Возвращает (число, ниже_порога: bool) или (None, None).
    """
    if raw is None:
        return None, None
    s = str(raw).lower().replace(" ", "").replace(",", ".")
    if s in ("", "-", "н/о", "но", "–"):
        return None, None
    below = "менее" in s
    m = re.search(r"(\d+(?:\.\d+)?)х?10\*?(\d+)", s)
    if not m:
        m2 = re.search(r"(\d+(?:\.\d+)?)", s)
        return (float(m2.group(1)), below) if m2 else (None, None)
    mant, exp = float(m.group(1)), int(m.group(2))
    return mant * (10 ** exp), below


def classify(indicator, raw):
    """Статус ячейки: 'несоответствие' / 'превышение' / 'норма' / 'не тестировали'."""
    s = _norm(raw).lower()
    if s in ("", "-", "–"):
        return "не тестировали"
    if "обнаруж" in s or s in ("рост", "+"):
        return "несоответствие"
    if indicator == "КМАФАнМ":
        val, below = parse_kma(raw)
        if val is not None and not below and val > KMA_LIMIT:
            return "превышение"
    return "норма"


def parse_workbook(path):
    """Читает Excel и возвращает длинную (tidy) таблицу: одна строка = один показатель одной пробы."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        if ws.title.strip().lower() in SKIP_SHEETS:
            continue
        headers = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        date_col = loc_col = None
        ind_cols = {}
        for i, h in enumerate(headers, start=1):
            hl = h.lower()
            if date_col is None and "дата" in hl:
                date_col = i
            if loc_col is None and "место" in hl:
                loc_col = i
            for ind, syns in INDICATORS.items():
                if ind in ind_cols:
                    continue
                if any(sy in hl for sy in syns):
                    ind_cols[ind] = i
                    break
        if not date_col or not loc_col or not ind_cols:
            continue  # это не журнал цеха

        cur_date = None
        for r in range(2, ws.max_row + 1):
            dv = ws.cell(r, date_col).value
            if dv not in (None, ""):
                cur_date = dv
            loc = _clean_loc(_norm(ws.cell(r, loc_col).value))
            cells = {ind: _norm(ws.cell(r, ci).value) for ind, ci in ind_cols.items()}
            if not loc and all(v in ("", "-") for v in cells.values()):
                continue
            if all(v in ("", "-") for v in cells.values()):
                continue  # пустая строка данных
            for ind, raw in cells.items():
                status = classify(ind, raw)
                kma_val, kma_below = (parse_kma(raw) if ind == "КМАФАнМ" else (None, None))
                rows.append({
                    "цех": ws.title.strip(),
                    "дата": pd.to_datetime(cur_date) if cur_date else pd.NaT,
                    "точка": loc,
                    "показатель": ind,
                    "значение": raw,
                    "статус": status,
                    "кмафанм_кое": kma_val,
                })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df[df["точка"] != ""].reset_index(drop=True)
    return df


def to_sqlite(df, db_path="smyvy.db"):
    import sqlite3
    con = sqlite3.connect(db_path)
    out = df.copy()
    out["дата"] = out["дата"].dt.strftime("%Y-%m-%d")
    out.to_sql("smyvy", con, if_exists="replace", index=False)
    con.close()
    return db_path


if __name__ == "__main__":
    import sys
    src = sys.argv[1] if len(sys.argv) > 1 else "07. Результаты смывов 2026.xlsx"
    df = parse_workbook(src)
    print(f"Строк-показателей: {len(df)}")
    print(f"Проб (уник. цех+дата+точка): "
          f"{df.groupby(['цех','дата','точка']).ngroups}")
    print("Статусы:\n", df["статус"].value_counts())
    print("\nНесоответствия и превышения:")
    bad = df[df["статус"].isin(["несоответствие", "превышение"])]
    for _, r in bad.iterrows():
        d = r["дата"].strftime("%d.%m.%Y") if pd.notna(r["дата"]) else "—"
        print(f"  {d} | {r['цех']} | {r['точка'][:40]} | {r['показатель']} | {r['статус']}")
    to_sqlite(df)
    print("\nБаза сохранена: smyvy.db")
